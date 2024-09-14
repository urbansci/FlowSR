# =================================================================================================================
# Description: This script is used to evaluate the performance of existing models. 
# The dataset to use is specifed by `dataset` and `level` variables. 
# The models to evaluate are specified using `model_list`. 
# The script will output the evaluation results in an Excel file.
# =================================================================================================================
from sklearn.metrics import mean_squared_error, mean_absolute_error, mean_absolute_percentage_error
from numpy import exp, round
from lib_loaddata import *
from matplotlib import pyplot as plt
from scipy import optimize
from math import sqrt
import openpyxl

dataset = 'US'  # ["england", "US", "JJJ","gd_commute","gd_mobility"]
level = 'county'  # ["mlad", "county", "state", "csa","subdistrict"]
model_list = ["GM_Zipf", "GM_Pow","GM_Exp","RM","ERM","IO","OPS"]  # ["GM_Zipf", "GM_Pow","GM_Exp","RM","ERM","IO","OPS"]
use_work_attr = True

if dataset == 'gd_commute':
    flow, dist, iores, iowork, attr = load_gd_commute_data(select_feat=['home_pop','work_pop'], level=level)
elif dataset == 'gd_mobility':
    flow, dist, iores, attr = load_gd_mobility_data(select_feat=['pop'], level=level)
if dataset == 'england':
    flow, dist, iores, iowork, attr = load_england_data_files(level=level,
                select_feat=['respop', 'workpop'])
elif dataset == 'US':
    flow, dist, iores, iowork, attr = load_us_data_files(level=level,
                select_feat=['respop', 'workpop'])
elif dataset == 'JJJ':
    flow, dist, iores, attr = load_jjj_data_files(level=level, select_feat=['pop_wan'])
else:
    raise NotImplementedError
print(len(dist.keys()))

reswb = openpyxl.Workbook()
ws = reswb.active

def prepare_data(ws, use_work_attr):
    if dataset in ["england", "US","gd_commute"] and use_work_attr:
        fid = 1  # 0: res 1: work
        io = iowork
    else:
        fid = 0
        io = iores
    X = []
    Y = []
    ori_sep = [0]
    ws.append(["ori", "dest", "dist", "io", "dpop", "opop", "vol"])
    for o in dist.keys():
        for d in flow[o].keys():
            X.append([dist[o][d], io[o][d], attr[d][fid], attr[o][fid]])
            Y.append(flow[o][d])
            ws.append([o, d, dist[o][d], io[o][d], attr[d][fid], attr[o][fid], flow[o][d]])
        ori_sep.append(ori_sep[-1]+len(flow[o].keys()))
    Xarr = np.asarray(X)
    Yarr = np.asarray(Y)
    outflow = [sum(flow[o].values()) for o in dist.keys()]
    return Xarr, Yarr, outflow, ori_sep

# Prepare Data for running models
Xarr, Yarr, outflow, ori_sep = prepare_data(ws, use_work_attr)
print(len(Yarr))

def allocation_law(dis, io, md, mo, param=None):
    if type =="GM_Zipf":
        p = md/dis
    elif type == "GM_Pow":
        p = md/(dis**param)
    elif type == "GM_Exp":
        p = md/exp(param*dis)
    elif type == "RM":
        p = md/((mo+io)*(mo+io+md))
    elif type == "ERM":
        p = ((mo+io+md)**param-(mo+io)**param)*(1+mo**param)/\
            ((1+(mo+io)**param)*(1+(mo+io+md)**param))
    elif type == "IO":
        p = exp(param*io)-exp(param*(io+md))
    elif type == "OPS":
        p = md/(mo+io+md)
    elif type == "custom":
        p = (md ** 0.5127566392825407) / (dis ** 3 / md + mo)
    else:
        raise NotImplementedError
    return p


def mse_loss(param):
    alloc = allocation_law(Xarr[:, 0], Xarr[:, 1], Xarr[:,2], Xarr[:,3], param)
    for oid in range(len(dist.keys())):
        stdfac = sum(alloc[ori_sep[oid]: ori_sep[oid+1]])
        alloc[ori_sep[oid]: ori_sep[oid+1]] *= outflow[oid]/stdfac
    return mean_squared_error(Yarr, alloc)


def pred(param=None):
    alloc = allocation_law(Xarr[:, 0], Xarr[:, 1], Xarr[:, 2], Xarr[:, 3], param)
    for oid in range(len(dist.keys())):
        stdfac = sum(alloc[ori_sep[oid]: ori_sep[oid + 1]])
        alloc[ori_sep[oid]: ori_sep[oid + 1]] *= outflow[oid] / stdfac
    return alloc

mid = 0
for type in model_list:
    if type in ["GM_Pow", "GM_Exp", "IO","ERM"]:
        if type == "IO":
            init_param = [-0.001]  # Need Change
            bound = [(-0.15, -0.0001)]  # IO: [(0.001, 1)]
            met = "Nelder-Mead"  # ["BFGS","Nelder-Mead"]
        else:
            init_param = [1.0]  # Need Change
            bound = None  # IO: [(0.001, 1)]
            met = "Nelder-Mead"
        #print(mse_loss(init_param))
        res = optimize.minimize(mse_loss, np.asarray(init_param), method=met, options={"disp": True},
                                bounds=bound)
        print(res.x, res.fun)
        Ypred = pred(res.x)
    else:  # parameter-free model
        Ypred = pred()

    rmse = sqrt(mean_squared_error(Yarr, Ypred))
    mae = mean_absolute_error(Yarr, Ypred)
    mape = mean_absolute_percentage_error(Yarr, Ypred)
    cpc = 1-sum(np.abs(Ypred-Yarr))/(sum(Yarr)+sum(Ypred))
    print(rmse, mae, mape, cpc)

    # Output Predicted Flow
    ws.cell(row=1, column=8+mid).value = type
    for i in range(len(Ypred)):
        ws.cell(row=2+i, column=8+mid).value = Ypred[i]

    plt.figure(figsize=(6, 6))
    plt.loglog(Yarr, Ypred, '.', markersize=1)
    u = np.linspace(1, max(max(Yarr), max(Ypred)), 5000)
    plt.loglog(u, u)
    plt.xlim((0.8, 1.1*max(max(Yarr), max(Ypred))))
    plt.ylim((0.8, 1.1*max(max(Yarr), max(Ypred))))
    plt.xlabel('Truth')
    plt.ylabel('Prediction')
    plt.show()

    mid += 1

reswb.save("bench_"+ dataset + "_" + level + ".xlsx")
