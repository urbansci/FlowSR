# =================================================================================================================
# Description: This file contains the functions to load the data for existing models' evaluation. 
# The functions in this file will be called in the ./bench_allocation.py, depending on the dataset used for evaluation.
# =================================================================================================================
import openpyxl
import numpy as np
import pandas as pd
import pickle
from tqdm import tqdm

def load_england_data_files(level='mlad', select_feat=None, modified_io=False):
    # feat: dist, o, d
    if level == 'msoa':
        feat = {'area_km2': 3, 'respop': 4, 'employedpop': 5, 'workpop': 6, 'households':7,
            'fb_pct': 8, 'deprived_pct': 9, 'nonwhite_pct': 10, 'bach_pct': 11, 'highsc_pct': 12}
    elif level == 'mlad':
        feat = {'respop': 4, 'workpop': 5}
    flow_file = open(f"../data/England/England_{level}_census11_supp3.pkl", 'rb')
    flow_dict = pickle.load(flow_file)
    dist_file = open(f"../data/England/England_{level}_dist.pkl", 'rb')
    dist_dict = pickle.load(dist_file)

    units = list(dist_dict.keys())
    geoid2row = dict()
    attrdata = openpyxl.load_workbook(f"../data/England/England_{level}_census11_attr.xlsx")
    attrtab = attrdata['attr']
    for r in range(2, attrtab.max_row + 1):
        geoid2row[int(attrtab.cell(r, 1).value[-6:])] = r
    attr_dict = dict()
    for o in units:
        if select_feat is None:
            attr = [attrtab.cell(geoid2row[o], c).value for c in feat.values()]
        else:
            attr = [attrtab.cell(geoid2row[o], feat[f]).value for f in select_feat]
        attr_dict[o] = attr

    if modified_io and level == 'mlad':
        iores_file = open(f"../data/England/England_{level}_miores.pkl", 'rb')
        iores_dict = pickle.load(iores_file)
        iowork_file = open(f"../data/England/England_{level}_miowork.pkl", 'rb')
        iowork_dict = pickle.load(iowork_file)
    else:
        iores_file = open(f"../data/England/England_{level}_iores.pkl", 'rb')
        iores_dict = pickle.load(iores_file)
        iowork_file = open(f"../data/England/England_{level}_iowork.pkl", 'rb')
        iowork_dict = pickle.load(iowork_file)
        if modified_io and level == 'msoa':
            for o in iores_dict.keys():
                for d in iores_dict[o].keys():
                    iores_dict[o][d] += attrtab.cell(geoid2row[o], feat["respop"]).value
                    iowork_dict[o][d] += attrtab.cell(geoid2row[o], feat["workpop"]).value
    return flow_dict, dist_dict, iores_dict, iowork_dict, attr_dict

def load_us_data_files(level='county', select_feat=None):
    # feat: dist, o, d
    if level == 'county':
        feat = {'respop': 4, 'employedpop': 5, 'workpop': 6}
    else:
        raise NotImplementedError

    flow_file = open(f"../data/US/us_acs15_{level}_flow.pkl", 'rb')
    flow_dict = pickle.load(flow_file)
    dist_file = open(f"../data/US/us_{level}_dist.pkl", 'rb')
    dist_dict = pickle.load(dist_file)
    iores_file = open(f"../data/US/us_{level}_iores.pkl", 'rb')
    iowork_file = open(f"../data/US/us_{level}_iowork.pkl", 'rb')
    iores_dict = pickle.load(iores_file)
    iowork_dict = pickle.load(iowork_file)
    msoa_units = list(dist_dict.keys())
    geoid2row = dict()
    attrdata = openpyxl.load_workbook(f"../data/US/us_acs15_{level}_attr.xlsx")
    attrtab = attrdata['attr']
    for r in range(2, attrtab.max_row + 1):
        geoid2row[int(attrtab.cell(r, 1).value)] = r
    attr_dict = dict()
    for o in msoa_units:
        if select_feat is None:
            attr = [attrtab.cell(geoid2row[o], c).value for c in feat.values()]
        else:
            attr = [attrtab.cell(geoid2row[o], feat[f]).value for f in select_feat]
        attr_dict[o] = attr
    return flow_dict, dist_dict, iores_dict, iowork_dict, attr_dict

def load_jjj_data_files(level='county', select_feat=None):
    if level == 'county':
        feat = {'area_km2': 8, 'pop_wan': 9, 'gdp_yi': 10}
    else:
        raise NotImplementedError
    flow_file = open(f"../data/Jingjinji/JJJ_{level}_flow.pkl", 'rb')
    flow_dict = pickle.load(flow_file)
    dist_file = open(f"../data/Jingjinji/JJJ_{level}_dist.pkl", 'rb')
    dist_dict = pickle.load(dist_file)
    io_file = open(f"../data/Jingjinji/JJJ_{level}_io.pkl", 'rb')
    io_dict = pickle.load(io_file)
    units = list(dist_dict.keys())
    geoid2row = dict()
    attrdata = openpyxl.load_workbook(f"../data/Jingjinji/JJJ_{level}_attr.xlsx")
    attrtab = attrdata['attr']
    for r in range(2, attrtab.max_row + 1):
        geoid2row[int(attrtab.cell(r, 5).value)] = r
    attr_dict = dict()
    for o in units:
        if select_feat is None:
            attr = [attrtab.cell(geoid2row[o], c).value for c in feat.values()]
        else:
            attr = [attrtab.cell(geoid2row[o], feat[f]).value for f in select_feat]
        attr_dict[o] = attr
    return flow_dict, dist_dict, io_dict, attr_dict

def load_gd_commute_data(select_feat=None, level='subdistrict'):
    import numpy as np
    import pandas as pd
    import pickle

    # Load the raw data
    flow_array = np.load(r"..\GD_data\Commuting_{}\gd_commute_flow_matrix_inter{}.npy".format(level,level))
    with open(r"..\GD_data\Commuting_{}\gd_commute_ids_mapping_inter{}.pkl".format(level,level), "rb") as file:
        id_dict = pickle.load(file)
    dist_df = pd.read_csv(r"..\GD_data\Commuting_{}\gd_commute_dist_inter{}.csv".format(level,level))
    attr_df = pd.read_csv(r"..\GD_data\Commuting_{}\gd_commute_attr_inter{}.csv".format(level,level))
    oppo_df_res = pd.read_csv(r"..\GD_data\Commuting_{}\gd_commute_opportunity_inter{}_res.csv".format(level,level))
    oppo_df_work = pd.read_csv(r"..\GD_data\Commuting_{}\gd_commute_opportunity_inter{}_work.csv".format(level,level))
    # Check whether the selected features exist
    if select_feat and not set(select_feat).issubset(attr_df.columns):
        missing_feats = set(select_feat) - set(attr_df.columns)
        raise ValueError(f"Features {missing_feats} not found in the dataset")

    # Transform the DataFrame to a more efficient lookup structure
    pivot_df = dist_df.pivot(index='o_id', columns='d_id', values='geodesic_dist')
    dist_dict = pivot_df.to_dict(orient='index')
    print("======> dist dict loaded")
    pivot_df = oppo_df_res.pivot(index='o_id', columns='d_id', values='opportunity')
    oppo_dict_res = pivot_df.to_dict(orient='index')
    print("======> oppo dict res loaded")
    pivot_df = oppo_df_work.pivot(index='o_id', columns='d_id', values='opportunity')
    oppo_dict_work = pivot_df.to_dict(orient='index')
    print("======> oppo dict work loaded")
    # Initialize the dictionaries
    flow_dict = {o_id: {} for o_id in id_dict.values()}
    attr_dict = {}

    for i in range(flow_array.shape[0]):
        o_id = id_dict[i]

        # Specify id column based on the level
        if level == 'subdistrict':
            id_col = 'street_num'
        elif level == "county":
            id_col = "county"
        
        if select_feat is None:
            attr_dict[o_id] = attr_df[attr_df[id_col] == o_id].iloc[0].values[1:]
        else:
            attr_dict[o_id] = attr_df[attr_df[id_col] == o_id].iloc[0][select_feat].values
        
        # Drop the origin if there is no flow
        if flow_array[i].sum() == 0:
            flow_dict.pop(o_id)
            continue
        for j in range(flow_array.shape[1]):
            d_id = id_dict[j]
            if i != j and flow_array[i,j]!=0:  # Exclude the flow from the origin to itself
                flow_dict[o_id][d_id] = flow_array[i, j]
    return flow_dict, dist_dict, oppo_dict_res, oppo_dict_work, attr_dict

def load_gd_mobility_data(select_feat=None, level='subdistrict'):
    import numpy as np
    import pandas as pd
    import pickle

    # Load the raw data
    flow_array = np.load(r"..\GD_data\Mobility_{}\gd_mobility_flow_matrix_inter{}.npy".format(level,level))
    with open(r"..\GD_data\Mobility_{}\gd_mobility_ids_mapping_inter{}.pkl".format(level,level), "rb") as file:
        id_dict = pickle.load(file)
    dist_df = pd.read_csv(r"..\GD_data\Mobility_{}\gd_mobility_dist_inter{}.csv".format(level,level))
    attr_df = pd.read_csv(r"..\GD_data\Mobility_{}\gd_mobility_attr_inter{}.csv".format(level,level))
    oppo_df = pd.read_csv(r"..\GD_data\Mobility_{}\gd_mobility_opportunity_inter{}.csv".format(level,level))
    
    # Check whether the selected features exist
    if select_feat and not set(select_feat).issubset(attr_df.columns):
        missing_feats = set(select_feat) - set(attr_df.columns)
        raise ValueError(f"Features {missing_feats} not found in the dataset")

    # Transform the DataFrame to a more efficient lookup structure
    pivot_df = dist_df.pivot(index='o_id', columns='d_id', values='geodesic_dist')
    dist_dict = pivot_df.to_dict(orient='index')
    print("======> dist dict loaded")
    pivot_df = oppo_df.pivot(index='o_id', columns='d_id', values='opportunity')
    oppo_dict = pivot_df.to_dict(orient='index')
    print("======> oppo dict loaded")
    
    # Initialize the dictionaries
    flow_dict = {o_id: {} for o_id in id_dict.values()}
    attr_dict = {}

    for i in range(flow_array.shape[0]):
        o_id = id_dict[i]
        
        # Specify id column based on the level
        if level == 'subdistrict':
            id_col = 'street_num'
        elif level == "county":
            id_col = "county"
        
        if select_feat is None:
            attr_dict[o_id] = attr_df[attr_df[id_col] == o_id].iloc[0].values[1:]
        else:
            attr_dict[o_id] = attr_df[attr_df[id_col] == o_id].iloc[0][select_feat].values
        
        if flow_array[i].sum() == 0:
            flow_dict.pop(o_id)
            continue
        for j in range(flow_array.shape[1]):
            d_id = id_dict[j]
            if i != j and flow_array[i,j]!=0:  # Exclude the flow from the origin to itself
                flow_dict[o_id][d_id] = flow_array[i, j]
                
    return flow_dict, dist_dict, oppo_dict, attr_dict
