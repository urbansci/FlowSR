"""
This script is used to generate flow based on different existing models with different types and levels of noise.
The existing model selected to generate flow is specified by parameter `modeltype`.
Parameter `noisetype` specifies the type of noise, and `sigma` determines the standard deviation of the
normal distribution noise introduced to the flow.
"""
from sklearn.metrics import mean_squared_error, mean_absolute_error, mean_absolute_percentage_error
from math import sqrt, exp
import numpy as np
from numpy.random import seed, randn
from matplotlib import pyplot as plt
import json
import pickle
import openpyxl

cur_seed = 1231
seed(cur_seed)
meta = dict()
meta["seed"] = cur_seed

dataset = 'england'
level = 'msoa'
modeltype = "RM"  # ["GMZipf", "GMPow","GMExp","RM","ERM","IO","OPS"]
use_work_attr = True
noisetype = 'mul'  # ["mul", "logadd"]
sigma = 0


def load_england_data_files(level='msoa', select_feat=None):
    # feat: dist, o, d
    if level == 'msoa':
        feat = {'area_km2': 3, 'respop': 4, 'employedpop': 5, 'workpop': 6, 'households':7,
            'fb_pct': 8, 'deprived_pct': 9, 'nonwhite_pct': 10, 'bach_pct': 11, 'highsc_pct': 12}
    elif level == 'mlad':
        feat = {'respop': 4, 'workpop': 5}
    flow_file = open(f"../../Data/England/England_{level}_census11_supp3.pkl", 'rb')
    flow_dict = pickle.load(flow_file)
    dist_file = open(f"../../Data/England/England_{level}_dist.pkl", 'rb')
    dist_dict = pickle.load(dist_file)
    iores_file = open(f"../../Data/England/England_{level}_iores.pkl", 'rb')
    iores_dict = pickle.load(iores_file)
    iores_file = None   # when use_work_attr = True, we don't need iores_dict
    iores_dict = None
    iowork_file = open(f"../../Data/England/England_{level}_iowork.pkl", 'rb')
    iowork_dict = pickle.load(iowork_file)
    msoa_units = list(dist_dict.keys())
    geoid2row = dict()
    attrdata = openpyxl.load_workbook(f"../../Data/England/England_{level}_census11_attr.xlsx")
    attrtab = attrdata['attr']
    for r in range(2, attrtab.max_row + 1):
        geoid2row[int(attrtab.cell(r, 1).value[-6:])] = r
    attr_dict = dict()
    for o in msoa_units:
        if select_feat is None:
            attr = [attrtab.cell(geoid2row[o], c).value for c in feat.values()]
        else:
            attr = [attrtab.cell(geoid2row[o], feat[f]).value for f in select_feat]
        attr_dict[o] = attr
    return flow_dict, dist_dict, iores_dict, iowork_dict, attr_dict


if dataset == 'england':
    flow, dist, iores, iowork, attr = load_england_data_files(level=level,
                                                              select_feat=['respop', 'workpop'])
    param_dict = {'msoa': {"GMPow": 1.33317602, "GMExp": 0.18416155}}
else:
    raise NotImplementedError
units = sorted(dist.keys())
N = len(units)
print(N)

outflow = {resid: sum(flow[resid].values()) for resid in units}
if use_work_attr:
    fid = 1  # 0: res 1: work
    io = iowork
else:
    fid = 0
    io = iores
X = dict()
for resid in units:
    Xo = []
    for workid in units:
        if workid == resid:
            continue
        Xo.append([dist[resid][workid], io[resid][workid], attr[workid][fid], attr[resid][fid]])
    Xo = np.asarray(Xo)
    X[resid] = Xo


def allocation_law(dis, io, md, mo, param=None):
    if modeltype == "GMZipf":
        p = md / dis
    elif modeltype == "GMPow":
        p = md / (dis ** param)
    elif modeltype == "GMExp":
        p = md / exp(param * dis)
    elif modeltype == "RM":
        p = md / ((mo + io) * (mo + io + md))
    elif modeltype == "ERM":
        p = ((mo + io + md) ** param - (mo + io) ** param) * (1 + mo ** param) / \
            ((1 + (mo + io) ** param) * (1 + (mo + io + md) ** param))
    elif modeltype == "IO":
        p = exp(param * io) - exp(param * (io + md))
    elif modeltype == "OPS":
        p = md / (mo + io + md)
    else:
        raise NotImplementedError
    return p


def pred(param=None):
    Y = []
    for resid in units:
        Xo = X[resid]
        alloc = allocation_law(Xo[:, 0], Xo[:, 1], Xo[:, 2], Xo[:, 3], param)
        stdfac = sum(alloc)
        alloc *= outflow[resid] / stdfac
        Y = np.concatenate((Y, alloc))
    assert len(Y) == N * (N - 1)
    return np.asarray(Y)


if modeltype in ["GMPow", "GMExp", "IO", "ERM"]:
    Ymodel = pred(param_dict[level][modeltype])
    meta["param"] = param_dict[level][modeltype]
else:  # parameter-free model
    Ymodel = pred()

# Add Random Noise
if noisetype == 'mul':
    Noise = np.asarray([1 + sigma * randn() for f in range(len(Ymodel))])
    Yturb = Ymodel * Noise
elif noisetype == 'logadd':
    Noise = np.asarray([sigma * randn() for f in range(len(Ymodel))])
    Yturb = Ymodel * np.exp(Noise)
else:
    raise NotImplementedError
Ysyn = np.round(Yturb)

rmse = sqrt(mean_squared_error(Ymodel, Ysyn))
mae = mean_absolute_error(Ymodel, Ysyn)
mape = mean_absolute_percentage_error(Ymodel, Ysyn)
cpc = 1 - sum(np.abs(Ymodel - Ysyn)) / (sum(Ymodel) + sum(Ysyn))
meta['RMSE'], meta['MAE'], meta['MAPE'], meta['CPC'] = rmse, mae, mape, cpc

# Flow data
flowdict = {resid: dict() for resid in units}
flowhist = dict()

thres = 3
nflow = 0
sumflow = 0
for res in range(N):
    resid = units[res]
    for work in range(N):
        if work == res:
            continue
        workid = units[work]
        flowind = res * (N - 1) + work
        if work > res:
            flowind -= 1
        vol = int(Ysyn[flowind])
        if vol < thres:
            continue
        nflow += 1
        sumflow += vol
        flowdict[resid][workid] = vol
        if vol in flowhist.keys():
            flowhist[vol] += 1
        else:
            flowhist[vol] = 1
meta["flownum"], meta["flowsum"], meta["flowavg"] = nflow, sumflow, sumflow / nflow
meta["flowmax"] = max(flowhist.keys())
outdeg = [len(d) for d in flowdict.values()]
meta["degavg"], meta["degmax"], meta["degmin"] = sum(outdeg) / len(outdeg), max(outdeg), min(outdeg)

flowfile = open(f"../../Data/synthetic/England/engmsoa_{modeltype}_{noisetype}{sigma}_supp{thres}_{cur_seed}.pkl", "wb")
pickle.dump(flowdict, flowfile)
flowfile.close()

val = sorted(list(flowhist.keys()))
freq = [flowhist[v] for v in val]
plt.loglog(val, freq, '.', markersize=2)
plt.xlabel("Commuting Flow")
plt.ylabel("Frequency")
plt.savefig(f"../../Data/synthetic/England/engmsoa_{modeltype}_{noisetype}{sigma}_supp{thres}_{cur_seed}_meta.png")
meta["flowhist"] = flowhist

fmeta = open(f"../../Data/synthetic/England/engmsoa_{modeltype}_{noisetype}{sigma}_supp{thres}_{cur_seed}_meta.txt","w")
json.dump(meta, fmeta)
fmeta.close()
